import argparse
import os
import os.path as osp
import sys
import random
import torch
from sklearn.model_selection import KFold
from torch_geometric.data import Data, HeteroData

from openpyxl.reader.excel import load_workbook
sys.setrecursionlimit(1000000000)
sys.path.append(os.path.realpath('.'))

from src.classes import LncRNA, RNA2RNA
from src.classes import Protein
from src.classes import LncRNA_Protein_Interaction

def parse_args():
    parser = argparse.ArgumentParser(description="generate_dataset.")
    parser.add_argument('--projectName', help='project name')
    parser.add_argument('--interactionDatasetName', default="NPInter2", help='raw interactions dataset')
    parser.add_argument('--createBalanceDataset', default=1, type=int, help='Create a Balance dataset')
    #parser.add_argument('--reduce', default=0, type=int,
    #                    help='randomly reduce the source database, and also maintain one connected component')
    #parser.add_argument('--reduceRatio', default=0.5, help='reduce Ratio')
    parser.add_argument('--output', default=1, type=int, help='output dataset or not')

    args = parser.parse_args()
    return args


def read_interaction_dataset(dataset_path, dataset_name):
    interaction_list = []
    negative_interaction_list = []
    lncRNA_list = []
    protein_list = []
    lncRNA_name_index_dict = {}
    protein_name_index_dict = {}
    set_interactionKey = set()
    set_negativeInteractionKey = set()
    # 列表：正负相互作用，lncRNA，蛋白质
    # 字典：rna-id，蛋白质-id
    # 集合：正负相互作用
    if not osp.exists(dataset_path):
        raise FileNotFoundError("The dataset path does not exist.")
    wb = load_workbook(dataset_path)
    sheets = wb.worksheets
    sheet = sheets[0]
    rows = sheet.rows

    # 分别为 lncRNA 和蛋白质设置独立的 serial_number 计数器
    lncRNA_serial_number = 0
    protein_serial_number = 0
    flag = 0

    for row in rows:
        if flag == 0:
            flag = flag + 1
            continue
        # 读出这一行的每个元素,每一行对应一个interaction实例，如果这个interaction对应的lncRNA和protein还没创建，就创建它
        # 并在索引词典中加入它在lncRNA_list或者protein_list中的索引
        [lncRNA_name, protein_name, label] = [col.value for col in row]
        label = int(label)
        if lncRNA_name not in lncRNA_name_index_dict:  # 新的，没创建过的lncRNA
            temp_lncRNA = LncRNA(lncRNA_name, lncRNA_serial_number, 'LncRNA')
            lncRNA_list.append(temp_lncRNA)
            lncRNA_name_index_dict[lncRNA_name] = lncRNA_serial_number
            lncRNA_serial_number = lncRNA_serial_number + 1
        else:  # 在interaction dataset中已经读到过，已经创建了对象的lncRNA，就存在lncRNA_list中
            temp_lncRNA = lncRNA_list[lncRNA_name_index_dict[lncRNA_name]]
        if protein_name not in protein_name_index_dict:  # 新的，没创建过的protein
            temp_protein = Protein(protein_name, protein_serial_number, 'Protein')
            protein_list.append(temp_protein)
            protein_name_index_dict[protein_name] = protein_serial_number
            protein_serial_number = protein_serial_number + 1
        else:  # 在interaction dataset中已经读到过，已经创建了对象的protein，就存在protein_list中
            temp_protein = protein_list[protein_name_index_dict[protein_name]]

        # 创建新的相互作用类型
        interaction_key = (temp_lncRNA.serial_number, temp_protein.serial_number)
        temp_interaction = LncRNA_Protein_Interaction(temp_lncRNA, temp_protein, label, interaction_key)
        temp_lncRNA.interaction_list.append(temp_interaction)
        temp_protein.interaction_list.append(temp_interaction)

        # 确认是正样本还是负样本
        if label == 1:
            interaction_list.append(temp_interaction)
            set_interactionKey.add(interaction_key)
        elif label == 0:
            negative_interaction_list.append(temp_interaction)
            set_negativeInteractionKey.add(interaction_key)
        else:
            print(label, 'is not supported.')

    print('number of lncRNA：{:d}, number of protein：{:d}, number of node：{:d}'.format(lncRNA_serial_number, protein_serial_number,
                                                                                      lncRNA_serial_number + protein_serial_number))
    print('number of interaction：{:d}'.format(len(interaction_list) + len(negative_interaction_list)))
    return interaction_list, negative_interaction_list, lncRNA_list, protein_list, lncRNA_name_index_dict, \
        protein_name_index_dict, set_interactionKey, set_negativeInteractionKey

def read_kmer_dataset(rna_list, protein_list):
    global args
    rna_kmer_path = 'data/lncRNA_3_mer/'+ args.interactionDatasetName +'/lncRNA_3_mer.txt'
    protein_kmer_path = 'data/protein_2_mer/'+ args.interactionDatasetName +'/protein_2_mer.txt'

    # 读取 RNA 的 k - mer 数据
    rna_kmer_dict = {}
    with open(rna_kmer_path, 'r') as rna_file:
        lines = rna_file.readlines()
        current_name = None
        for line in lines:
            line = line.strip()
            if line.startswith('>'):
                current_name = line[1:]
            elif current_name:
                vector = [float(val) for val in line.split('\t')]
                rna_kmer_dict[current_name] = vector

    # 将 RNA 的 k - mer 数据存入 rna_list 的 embedded_vector
    for rna in rna_list:
        if rna.name in rna_kmer_dict:
            rna.embedded_vector = rna_kmer_dict[rna.name]

    # 读取蛋白质的 k - mer 数据
    protein_kmer_dict = {}
    with open(protein_kmer_path, 'r') as protein_file:
        lines = protein_file.readlines()
        current_name = None
        for line in lines:
            line = line.strip()
            if line.startswith('>'):
                current_name = line[1:]
            elif current_name:
                vector = [float(val) for val in line.split('\t')]
                protein_kmer_dict[current_name] = vector

    # 将蛋白质的 k - mer 数据存入 protein_list 的 embedded_vector
    for protein in protein_list:
        if protein.name in protein_kmer_dict:
            protein.embedded_vector = protein_kmer_dict[protein.name]

    return rna_list, protein_list


def negative_interaction_generation():
    global lncRNA_list, protein_list, interaction_list, negative_interaction_list, set_interactionKey, set_negativeInteractionKey
    set_negativeInteractionKey = set()
    if len(negative_interaction_list) != 0:
        raise Exception('negative interactions exist')

    num_of_interaction = len(interaction_list)
    num_of_lncRNA = len(lncRNA_list)
    num_of_protein = len(protein_list)
    negative_interaction_count = 0
    while (negative_interaction_count < num_of_interaction):
        random_index_lncRNA = random.randint(0, num_of_lncRNA - 1)
        random_index_protein = random.randint(0, num_of_protein - 1)
        temp_lncRNA = lncRNA_list[random_index_lncRNA]
        temp_protein = protein_list[random_index_protein]
        # 检查随机选出的lncRNA和protein是不是有已知相互作用
        key_negativeInteraction = (temp_lncRNA.serial_number, temp_protein.serial_number)
        if key_negativeInteraction in set_interactionKey:
            continue
        if key_negativeInteraction in set_negativeInteractionKey:
            continue

        # 经过检查，随机选出的lncRNA和protein是可以作为负样本的
        set_negativeInteractionKey.add(key_negativeInteraction)
        temp_interaction = LncRNA_Protein_Interaction(temp_lncRNA, temp_protein, 0, key_negativeInteraction)
        negative_interaction_list.append(temp_interaction)
        temp_lncRNA.interaction_list.append(temp_interaction)
        temp_protein.interaction_list.append(temp_interaction)
        negative_interaction_count = negative_interaction_count + 1
    print('generate ', len(negative_interaction_list), ' negative samples')


def read_jaccard_dataset(lncRNA_list, protein_list):
    lncRNA_Jaccard_dic = {}
    protein_Jaccard_dic = {}
    jaccard_list = []
    jaccard_number = 0
    # 获取lncRNA相互作用的蛋白质集合
    for lncRNA in lncRNA_list:
        protein_set = set()
        for interaction in lncRNA.interaction_list:
            if interaction.y == 1:
                protein_set.add(interaction.protein.serial_number)
        lncRNA_Jaccard_dic[lncRNA] = protein_set

    for protein in protein_list:
        rna_set = set()
        for interaction in protein.interaction_list:
            if interaction.y == 1:
                rna_set.add(interaction.lncRNA.serial_number)
        protein_Jaccard_dic[protein] = rna_set

    # 计算lncRNA两两之间的jaccard系数
    for i in range(len(lncRNA_list) - 1):
        set1 = lncRNA_Jaccard_dic[lncRNA_list[i]]
        for j in range(i + 1, len(lncRNA_list)):
            set2 = lncRNA_Jaccard_dic[lncRNA_list[j]]
            jaccard = len(set1.intersection(set2)) / len(set1.union(set2))
            if jaccard > 0.8:
                jaccard_list.append(RNA2RNA(lncRNA_list[i], lncRNA_list[j], 1, jaccard_number))
                jaccard_number = jaccard_number + 1
                # print(lncRNA_list[i].name, lncRNA_list[j].name)
    r2r_num = jaccard_number
    # 计算蛋白质两两之间的jaccard系数
    for i in range(len(protein_list) - 1):
        set1 = protein_Jaccard_dic[protein_list[i]]
        for j in range(i + 1, len(protein_list)):
            set2 = protein_Jaccard_dic[protein_list[j]]
            jaccard = len(set1.intersection(set2)) / len(set1.union(set2))
            if jaccard > 0:
                jaccard_list.append(RNA2RNA(protein_list[i], protein_list[j], 0, jaccard_number))
                jaccard_number = jaccard_number + 1
                # print(protein_list[i].name, protein_list[j].name)
    print('number of jaccard RNA2RNA: {:d}'.format(r2r_num))
    print('number of jaccard P2P: {:d}'.format(jaccard_number-r2r_num))
    print('number of Jaccard: {:d}'.format(len(jaccard_list)))
    return jaccard_list


def read_blast_dataset(lncRNA_name_index_dict, protein_name_index_dict):
    global lncRNA_list, protein_list
    blast_list = []
    rna_pairs_path = 'rna_rna_pairs.xlsx'
    protein_pairs_path = 'protein_protein_pairs.xlsx'

    if not osp.exists(rna_pairs_path) or not osp.exists(protein_pairs_path):
        raise FileNotFoundError("One or both of the pair files do not exist.")
    # 读取rna_rna_pairs.xlsx
    wb_rna = load_workbook(rna_pairs_path)
    sheet_rna = wb_rna.worksheets[0]
    rows_rna = sheet_rna.rows
    cnt = 0
    for row in rows_rna:
        rna1_name, rna2_name = [col.value for col in row]
        if rna1_name in lncRNA_name_index_dict and rna2_name in lncRNA_name_index_dict:
            rna1 = lncRNA_list[lncRNA_name_index_dict[rna1_name]]
            rna2 = lncRNA_list[lncRNA_name_index_dict[rna2_name]]
            blast_list.append(RNA2RNA(rna1, rna2,1,cnt))
            rna_cnt = cnt + 1
    number = len(blast_list)
    cnt = 0
    # 读取protein_protein_pairs.xlsx
    wb_protein = load_workbook(protein_pairs_path)
    sheet_protein = wb_protein.worksheets[0]
    rows_protein = sheet_protein.rows
    for row in rows_protein:
        protein1_name, protein2_name = [col.value for col in row]
        if protein1_name in protein_name_index_dict and protein2_name in protein_name_index_dict:
            protein1 = protein_list[protein_name_index_dict[protein1_name]]
            protein2 = protein_list[protein_name_index_dict[protein2_name]]
            blast_list.append(RNA2RNA(protein1, protein2,0,cnt))
            cnt = cnt + 1
    print('number of blast RNA2RNA: {:d}'.format(number))
    print('number of blast P2P: {:d}'.format(len(blast_list) - number))
    print('number of blast: {:d}'.format(len(blast_list)))

    return blast_list


def create_pyg_graph_jaccard(interaction_list, lncRNA_list, protein_list, jaccard_list):
    # 初始化异质图数据
    graph = HeteroData()

    # 节点特征矩阵
    lncRNA_features = []
    protein_features = []

    # 将 RNA 的 embedded_vector 作为节点特征
    for rna in lncRNA_list:
        lncRNA_features.append(rna.embedded_vector)
    lncRNA_features = torch.tensor(lncRNA_features, dtype=torch.float)

    # 将蛋白质的 embedded_vector 作为节点特征
    for protein in protein_list:
        protein_features.append(protein.embedded_vector)
    protein_features = torch.tensor(protein_features, dtype=torch.float)

    # 添加节点特征
    graph['lncRNA'].x = lncRNA_features
    graph['protein'].x = protein_features

    # 获取相互作用边信息
    lncRNA_protein_edge_index = []
    for interaction in interaction_list:
        lncRNA_protein_edge_index.append([interaction.lncRNA.serial_number, interaction.protein.serial_number])
    lncRNA_protein_edge_index = torch.tensor(lncRNA_protein_edge_index, dtype=torch.long).t().contiguous()

    # 分离 rna-rna 和蛋白质 - 蛋白质的 jaccard 关联边
    rna_rna_edge_index = []
    protein_protein_edge_index = []
    for r2r_interaction in jaccard_list:
        if r2r_interaction.y == 1:  # RNA 关联
            rna_rna_edge_index.append([r2r_interaction.lncRNA1.serial_number, r2r_interaction.lncRNA2.serial_number])
        elif r2r_interaction.y == 0:  # 蛋白质关联
            protein_protein_edge_index.append([r2r_interaction.lncRNA1.serial_number, r2r_interaction.lncRNA2.serial_number])

    rna_rna_edge_index = torch.tensor(rna_rna_edge_index, dtype=torch.long).t().contiguous()
    protein_protein_edge_index = torch.tensor(protein_protein_edge_index, dtype=torch.long).t().contiguous()

    # 添加边信息
    graph['lncRNA', 'interacts_with', 'protein'].edge_index = lncRNA_protein_edge_index
    graph['lncRNA', 'jaccard_related', 'lncRNA'].edge_index = rna_rna_edge_index
    graph['protein', 'jaccard_related', 'protein'].edge_index = protein_protein_edge_index

    return graph

def create_pyg_graph_blast(interaction_list, lncRNA_list, protein_list, blast_list):
    # 初始化异质图数据
    graph = HeteroData()

    # 节点特征矩阵
    lncRNA_features = []
    protein_features = []

    # 将 RNA 的 embedded_vector 作为节点特征
    for rna in lncRNA_list:
        lncRNA_features.append(rna.embedded_vector)
    lncRNA_features = torch.tensor(lncRNA_features, dtype=torch.float)

    # 将蛋白质的 embedded_vector 作为节点特征
    for protein in protein_list:
        protein_features.append(protein.embedded_vector)
    protein_features = torch.tensor(protein_features, dtype=torch.float)

    # 添加节点特征
    graph['lncRNA'].x = lncRNA_features
    graph['protein'].x = protein_features

    # 获取相互作用边信息
    lncRNA_protein_edge_index = []
    for interaction in interaction_list:
        lncRNA_protein_edge_index.append([interaction.lncRNA.serial_number, interaction.protein.serial_number])
    lncRNA_protein_edge_index = torch.tensor(lncRNA_protein_edge_index, dtype=torch.long).t().contiguous()

    # 分离 rna-rna 和蛋白质 - 蛋白质的 blast 关联边
    rna_rna_edge_index = []
    protein_protein_edge_index = []
    for r2r_interaction in blast_list:
        if r2r_interaction.y == 1:  # RNA 关联
            rna_rna_edge_index.append([r2r_interaction.lncRNA1.serial_number, r2r_interaction.lncRNA2.serial_number])
        elif r2r_interaction.y == 0:  # 蛋白质关联
            protein_protein_edge_index.append([r2r_interaction.lncRNA1.serial_number, r2r_interaction.lncRNA2.serial_number])

    rna_rna_edge_index = torch.tensor(rna_rna_edge_index, dtype=torch.long).t().contiguous()
    protein_protein_edge_index = torch.tensor(protein_protein_edge_index, dtype=torch.long).t().contiguous()

    # 添加边信息
    graph['lncRNA', 'interacts_with', 'protein'].edge_index = lncRNA_protein_edge_index
    graph['lncRNA', 'blast_related', 'lncRNA'].edge_index = rna_rna_edge_index
    graph['protein', 'blast_related', 'protein'].edge_index = protein_protein_edge_index

    return graph


def generate_training_and_testing(graph_jaccard, graph_blast):
    global interaction_list, negative_interaction_list,args
    all_interactions = interaction_list + negative_interaction_list
    random.shuffle(all_interactions)

    # 分离正样本和负样本
    positive_samples = [interaction for interaction in all_interactions if interaction.y == 1]
    negative_samples = [interaction for interaction in all_interactions if interaction.y == 0]

    # 确保正负样本数量相等
    min_samples = min(len(positive_samples), len(negative_samples))
    positive_samples = positive_samples[:min_samples]
    negative_samples = negative_samples[:min_samples]

    all_samples = positive_samples + negative_samples
    random.shuffle(all_samples)

    kf = KFold(n_splits=5, shuffle=False)

    for fold, (train_index, test_index) in enumerate(kf.split(all_samples)):
        train_samples = [all_samples[i] for i in train_index]
        test_samples = [all_samples[i] for i in test_index]

        # 分离测试集中的正样本
        test_positive_samples = [sample for sample in test_samples if sample.y == 1]

        # 生成子图
        subgraph_jaccard = graph_jaccard.clone()
        subgraph_blast = graph_blast.clone()

        # 删除子图中测试集正样本的边
        for sample in test_positive_samples:
            lncRNA_index = sample.lncRNA.serial_number
            protein_index = sample.protein.serial_number

            edge_index = subgraph_jaccard['lncRNA', 'interacts_with', 'protein'].edge_index
            mask = ~((edge_index[0] == lncRNA_index) & (edge_index[1] == protein_index))
            subgraph_jaccard['lncRNA', 'interacts_with', 'protein'].edge_index = edge_index[:, mask]

            edge_index = subgraph_blast['lncRNA', 'interacts_with', 'protein'].edge_index
            mask = ~((edge_index[0] == lncRNA_index) & (edge_index[1] == protein_index))
            subgraph_blast['lncRNA', 'interacts_with', 'protein'].edge_index = edge_index[:, mask]

        # 保存训练集、测试集和子图
        fold_dir = 'data/graph/'+args.projectName+f'/fold_{fold}'
        os.makedirs(fold_dir, exist_ok=True)

        torch.save(train_samples, os.path.join(fold_dir, 'train_samples.pt'))
        torch.save(test_samples, os.path.join(fold_dir, 'test_samples.pt'))
        torch.save(subgraph_jaccard, os.path.join(fold_dir, 'subgraph_jaccard.pt'))
        torch.save(subgraph_blast, os.path.join(fold_dir, 'subgraph_blast.pt'))

        print(f"Fold {fold} saved successfully.")
        print(subgraph_jaccard)
        print(subgraph_blast)
        print('*******************************************')


if __name__ == '__main__':
    args = parse_args()
    interaction_dataset_path = 'data/source_database_data/' + args.interactionDatasetName + '.xlsx'
    interaction_list, negative_interaction_list, lncRNA_list, protein_list, lncRNA_name_index_dict, protein_name_index_dict, set_interactionKey, \
        set_negativeInteractionKey = read_interaction_dataset(dataset_path=interaction_dataset_path,
                                                              dataset_name=args.interactionDatasetName)
    if args.createBalanceDataset == 1:
        negative_interaction_generation() # 生成负样本
    lncRNA_list, protein_list = read_kmer_dataset(lncRNA_list,protein_list)
    jaccard_list = read_jaccard_dataset(lncRNA_list=lncRNA_list, protein_list=protein_list)
    blast_list = read_blast_dataset(lncRNA_name_index_dict=lncRNA_name_index_dict, protein_name_index_dict=protein_name_index_dict)

    graph_jaccard = create_pyg_graph_jaccard(interaction_list, lncRNA_list, protein_list, jaccard_list)
    graph_blast = create_pyg_graph_blast(interaction_list, lncRNA_list, protein_list, blast_list)
    generate_training_and_testing(graph_jaccard,graph_blast)
    print(graph_jaccard)
    print(graph_blast)